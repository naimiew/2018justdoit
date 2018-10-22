# -*- coding: utf-8 -*-
# 傻瓜式应付功能测试需要，批量提报附件生成器V0.1
#1.压缩包内容包含：1.营业执照照片，1张；2.开户许可证照片，1张；3.法人身份证正反面照片，各1张；
#2.照片命名规则：统一社会信用代码-类型
#类型可选值有 ：01 表示营业执照照片、02 表示开户许可证照片、
#03 表示身份证正面照片、04 表示身份证反面照片
#例如：
#xxxx-01，表示统一社会信用代码为 xxxx 的营业执照照片
#xxxx-02，表示统一社会信用代码为 xxxx 的开户许可证照片
#xxxx-03，表示统一社会信用代码为 xxxx 的身份证正面照片
#xxxx-04，表示统一社会信用代码为 xxxx 的身份证反面照片

import openpyxl
import shutil

# excel路径
excel_path = 'c:/Users/kte/Pictures/新建文件夹/'
excel_name = 'AAAAAA5.xlsx'
# jpg路径
old_jpg_path = 'c:/Users/kte/Pictures/新建文件夹/1022/'
new_jpg_path = 'c:/Users/kte/Pictures/新建文件夹/1000/'

# 打开文件：
rw_excel = openpyxl.load_workbook(excel_path + excel_name)
# 查找文件判断
if rw_excel != '':
	print('已找到excel文件，文件名为@@@@' + excel_name)
else:
	print('查找excel失败''！！！！！！！！')
# 获取sheet：
work_sheet = rw_excel.worksheets[0]  # 通过表名获取
print('该sheet页为' + work_sheet.title)

#  read_excel.get_sheet_by_name("Sheet2")  建议不适用这个过时的方法
# 获取行数和列数：
rows = work_sheet.max_row  # 获取行数
cols = work_sheet.max_column  # 获取列数
print('rows_num:' + str(rows) + '\n' + 'cols_num:' + str(cols))
# 设置启动行
go_row = 2
# 设置动态关联excel行列
while go_row <= rows:
	# 执行excel行
	go_row = go_row + 1
	# 获取单元格值：
	Data1 = work_sheet.cell(go_row, column=1).value
	Data14 = work_sheet.cell(go_row, column=14).value
	text_Data = str(Data1) + '##' + str(Data14)
	# 缺少处理空的data14的问题，也没有把目标文件和新文件做活
	old_jpg1_str = old_jpg_path + '1.jpg'
	old_jpg2_str = old_jpg_path + '2.jpg'
	old_jpg3_str = old_jpg_path + '3.jpg'
	old_jpg4_str = old_jpg_path + '4.jpg'
	new_jpg1_str = new_jpg_path + str(Data14) + '-01.jpg'
	new_jpg2_str = new_jpg_path + str(Data14) + '-02.jpg'
	new_jpg3_str = new_jpg_path + str(Data14) + '-03.jpg'
	new_jpg4_str = new_jpg_path + str(Data14) + '-04.jpg'
	
	shutil.copyfile(old_jpg1_str, new_jpg1_str)
	shutil.copyfile(old_jpg2_str, new_jpg2_str)
	shutil.copyfile(old_jpg3_str, new_jpg3_str)
	shutil.copyfile(old_jpg4_str, new_jpg4_str)
	
	print(text_Data)
	# print(Data)
	# 获取数据判断 if not 判空与真
	if Data14:
		print('获取数据成功！！！！！！！！')
	else:
		print('获取数据失败！！！！！！！！')

# 拼贴完整路径
excel_path_full = excel_path + excel_name
print(excel_path_full)
